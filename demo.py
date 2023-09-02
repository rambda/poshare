from datetime import date, datetime
import time
from pathlib import Path
import csv
from openpyxl import Workbook
from poshare import Xueqiu
from config import config


def json_to_pf(dict):
    dict['link'] = 'https://xueqiu.com/P/%s' % dict['symbol']
    dict['analysis_link'] = 'https://xueqiu.com/service/p/cube-analyze?symbol=%s' % dict['symbol']
    dict['owner_link'] = 'https://xueqiu.com/u/%s' % dict['owner_id']
    dict['annualized_gain_rate'] = dict['annualized_gain_rate'] / 100.0
    created_at = int(dict['created_at'] / 1000)
    dict['created_at'] = created_at
    dict['created_date'] = datetime.fromtimestamp(created_at).strftime('%Y-%m-%d')
    updated_at = int(dict['updated_at'] / 1000)
    dict['updated_at'] = updated_at
    dict['updated_date'] = datetime.fromtimestamp(created_at).strftime('%Y-%m-%d')
    closed_at = dict['closed_at']
    if closed_at is not None:
        closed_at = int(closed_at / 1000)
        duration = date.fromtimestamp(closed_at) - date.fromtimestamp(created_at)
        dict['duration_years'] = duration.days / 365.0
    else:
        duration = date.today() - date.fromtimestamp(created_at)
        dict['duration_years'] = duration.days / 365.0

    # filtered_dict = {key: dict[key] for key in dict
    #                  if key in {field.name for field in dataclasses.fields(Portfolio)}}
    #
    # return Portfolio(**filtered_dict)
    return dict


# for symbol in ["%06d" % i for i in range(0, 1e6) + "%06d" % i for i in range(1e6, 1e7)]:


Xueqiu.config(config)

fdir = Path("download")
# fdir.mkdir(parents=True, exist_ok=True)

meta_path = fdir / ("meta_" + time.strftime("%Y%m") + ".csv")
meta_path.touch()
excel_path = fdir / ("excel_" + time.strftime("%Y%m") + ".xlsx")
excel_path.touch()

workbook = Workbook()
sheet = workbook.active

fields = ['symbol', 'name', 'market', 'duration_years', 'annualized_gain_rate', 'net_value', 'analysis_link',
          'description', 'owner_id', 'created_date', "updated_date", 'closed_at']
sheet.append(fields)

symbols = [row[0] for row in csv.reader(meta_path.open('r', encoding='utf-8'))]

with meta_path.open('a', newline='\n', encoding='utf-8') as f:
    writer = csv.writer(f)

    # for symbol in ['ZH2363479', 'ZH1350829']:
    n1e6 = int(1e6)
    n1e7 = int(1e7)
    for symbol in ["ZH%06d" % i for i in range(0, n1e6)] + ["ZH%07d" % i for i in range(n1e6, n1e7)]:

        if symbol in symbols:
            # print("skipped.")
            continue

        xq = Xueqiu(symbol=symbol)
        time.sleep(1.5)

        if xq.bad_data:
            writer.writerow([symbol, None, None])
            print(symbol, " ", "bad data.", "\n")
            continue

        if xq.need_retry:
            writer.writerow([symbol + "_retry", None, None])
            print(symbol, " ", "need_retry.", "\n")
            continue

        pf = json_to_pf(xq.cube_info)

        writer.writerow([symbol, pf['annualized_gain_rate'], pf['duration_years']])
        print(symbol, " ", "CAGR=", pf['annualized_gain_rate'], " ", "YEAR=", pf['duration_years'], "\n")

        if pf['annualized_gain_rate'] > 0.5 and pf['net_value'] > 5.0:
            # print([pf[field] for field in fields])
            sheet.append([pf[field] for field in fields])
            # print(pf)
            name_cell = sheet.cell(row=sheet.max_row, column=fields.index('name') + 1)
            name_cell.hyperlink = pf['link']
            name_cell.style = "Hyperlink"

            analysis_link_cell = sheet.cell(row=sheet.max_row, column=fields.index('analysis_link') + 1)
            analysis_link_cell.hyperlink = pf['analysis_link']
            analysis_link_cell.hyperlink.display = "Analysis"
            analysis_link_cell.style = "Hyperlink"

            owner_id_cell = sheet.cell(row=sheet.max_row, column=fields.index('owner_id') + 1)
            owner_id_cell.hyperlink = pf['owner_link']
            owner_id_cell.style = "Hyperlink"

workbook.save(excel_path)
